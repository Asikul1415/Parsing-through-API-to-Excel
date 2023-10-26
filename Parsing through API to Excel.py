import requests as re
import json
import bs4 as BS4
from openpyxl import *

api_key = "YOUR_API_KEY"
wb = load_workbook("YOUR_DIRECTORY")
sheet = wb.get_sheet_by_name("Test")
inn = []
k = 2

while(sheet.cell(column=1,row=k).value != None):
    inn.append(sheet.cell(column=1,row=k).value)
    k += 1   
k = 2

for i in range(len(inn)):
    get = f"https://api.checko.ru/v2/company?key={api_key}&inn={inn[i]}"
    response = re.get(get, headers={'User-Agent': 'Mozilla/5.0'})
    response_json = json.loads(response.text)['data']

    try:
        sheet[f"B{k}"] = response_json['НаимСокр']
        sheet[f"C{k}"] = response_json['Руковод'][0]['ФИО']
        sheet[f"D{k}"] = response_json['ОГРН']
        sheet[f"E{k}"] = response_json['ОКПО']
        sheet[f"F{k}"] = response_json['КПП']
        sheet[f"G{k}"] = response_json['ОКВЭД']['Наим']
    except:
        k += 1
        continue
    
    k += 1

wb.save("test.xlsx")

